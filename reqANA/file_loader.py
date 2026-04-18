from __future__ import annotations

from io import BytesIO
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook


SUPPORTED_TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".tsv"}
SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
SUPPORTED_EXTENSIONS = SUPPORTED_TEXT_EXTENSIONS | SUPPORTED_EXCEL_EXTENSIONS


async def read_requirement_file(file: UploadFile) -> str:
    suffix = Path(file.filename or "").suffix.lower()
    raw = await file.read()
    return read_requirement_bytes(file.filename or "uploaded-file", raw)


def read_requirement_bytes(filename: str, raw: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported requirement file type '{suffix}'. "
            f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}."
        )

    if suffix in SUPPORTED_EXCEL_EXTENSIONS:
        return _read_excel(raw)
    return raw.decode("utf-8")


def _read_excel(raw: bytes) -> str:
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    sections: list[str] = []
    for worksheet in workbook.worksheets:
        rows: list[str] = []
        for row in worksheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if any(values):
                rows.append("\t".join(values).rstrip())
        if rows:
            sections.append(f"## Sheet: {worksheet.title}\n" + "\n".join(rows))
    return "\n\n".join(sections).strip()
